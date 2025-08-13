import os
import time
import requests
import pandas as pd
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from datetime import datetime, timedelta
from typing import List, Dict, Optional
from tqdm import tqdm
import threading
import re

class VirusTotalDomainReviewer:
    def __init__(self, api_key: str, account_type: str = "personal"):
        self.api_key = api_key
        self.account_type = account_type
        self.base_url = "https://www.virustotal.com/api/v3"
        self.headers = {
            "accept": "application/json",
            "x-apikey": api_key
        }
        # Set rate limit based on account type
        self.rate_limit_delay = 1 if account_type == "enterprise" else 15

    def defang_domain(self, domain: str) -> str:
        """
        Remove defang characters from domain (e.g., google(.)com -> google.com)
        """
        # Remove parentheses and brackets around dots
        domain = re.sub(r'\(\.\)', '.', domain)
        domain = re.sub(r'\[\.\]', '.', domain)
        
        # Remove any other parentheses or brackets
        domain = re.sub(r'[\(\)\[\]]', '', domain)
        
        return domain.strip()

    def calculate_risk_score(self, reputation_score, malicious_vendors, suspicious_vendors, categories, is_newly_registered):
        """
        Calculate risk score based on multiple factors
        """
        risk_score = 0
        is_high_risk = False
        
        # Check for high-risk conditions that force high risk categorization
        risk_words = ['malicious', 'phishing', 'abuse', 'scam', 'fraud', 'malware', 
                     'suspicious', 'dangerous', 'threat', 'attack', 'exploit']
        
        # High-risk condition 1: Category contains risk words
        has_risk_words = False
        if categories != 'N/A' and categories:
            categories_lower = categories.lower()
            for word in risk_words:
                if word in categories_lower:
                    has_risk_words = True
                    break
        
        # High-risk condition 2: Reputation score greater than -5
        has_bad_reputation = False
        if reputation_score != 'N/A' and reputation_score is not None:
            if reputation_score > -5 and reputation_score < 0:
                has_bad_reputation = True
        
        # High-risk condition 3: Newly registered domain
        is_newly_reg = bool(is_newly_registered)
        
        # If any high-risk condition is met, set as high risk
        if has_risk_words or has_bad_reputation or is_newly_reg:
            is_high_risk = True
        
        # Calculate base risk score
        # Reputation score (negative = higher risk)
        if reputation_score != 'N/A' and reputation_score is not None:
            if reputation_score < 0:
                risk_score += abs(reputation_score) * 0.1  # Scale down large negative values
            elif reputation_score == 0:
                risk_score += 5  # Neutral reputation
        
        # Malicious vendors count
        if malicious_vendors != 'N/A' and malicious_vendors is not None:
            risk_score += malicious_vendors * 10
        
        # Suspicious vendors count
        if suspicious_vendors != 'N/A' and suspicious_vendors is not None:
            risk_score += suspicious_vendors * 5
        
        # Categories containing risk words
        if has_risk_words:
            risk_score += 15
        
        # Newly registered domain
        if is_newly_registered:
            risk_score += 20
        
        # Force high risk score (minimum 71) if high-risk conditions are met
        if is_high_risk:
            risk_score = max(risk_score, 71)
        
        return min(risk_score, 100)  # Cap at 100

    def verify_api_key(self) -> bool:
        """
        Verify if the API key is valid by making a test request to VirusTotal
        """
        try:
            # Use a simple endpoint to test the API key
            url = f"{self.base_url}/users/current"
            response = requests.get(url, headers=self.headers)
            
            if response.status_code == 200:
                return True
            elif response.status_code == 401:
                return False  # Invalid API key
            else:
                # Other status codes might indicate network issues, but not necessarily invalid key
                return True
        except Exception as e:
            # Network errors or other exceptions
            return False

    def get_domain_info(self, domain: str) -> Optional[Dict]:
        """
        Query VirusTotal API for domain information
        """
        try:
            url = f"{self.base_url}/domains/{domain}"
            response = requests.get(url, headers=self.headers)
            if response.status_code == 200:
                return response.json()
            elif response.status_code == 404:
                return None
            elif response.status_code == 429:
                return None
            else:
                return None
        except Exception as e:
            return None

    def parse_domain_data(self, domain: str, data: Dict) -> Dict:
        try:
            attributes = data.get('data', {}).get('attributes', {})

            # Get reputation data
            reputation = attributes.get('reputation', 0)
            is_malicious = reputation < 0

            # Get last analysis stats
            last_analysis_stats = attributes.get('last_analysis_stats', {})
            malicious_count = last_analysis_stats.get('malicious', 0)
            suspicious_count = last_analysis_stats.get('suspicious', 0)
            harmless_count = last_analysis_stats.get('harmless', 0)
            undetected_count = last_analysis_stats.get('undetected', 0)

            # Get last analysis date
            last_analysis_date = attributes.get('last_analysis_date')
            if last_analysis_date:
                last_analysis_date = datetime.fromtimestamp(last_analysis_date).strftime('%Y-%m-%d %H:%M:%S')

            # Get tags
            tags = attributes.get('tags', [])
            tags_str = ', '.join(tags) if tags else 'N/A'

            # Get categories
            categories = attributes.get('categories', {})
            categories_str = ', '.join([f"{k}: {v}" for k, v in categories.items()]) if categories else 'N/A'

            # Get total vendors
            last_analysis_results = attributes.get('last_analysis_results', {})
            total_vendors = len(last_analysis_results) if last_analysis_results else 0

            # WHOIS/registration info from VirusTotal
            creation_date = attributes.get('creation_date')
            expiration_date = attributes.get('expiration_date')
            registrar = attributes.get('registrar', 'N/A')
            whois_status = 'Success' if creation_date or expiration_date or registrar != 'N/A' else 'Not available'

            # Parse and format dates
            creation_date_obj = None
            if creation_date:
                try:
                    creation_date_obj = datetime.fromtimestamp(creation_date)
                    creation_date_str = creation_date_obj.strftime('%Y-%m-%d')
                except Exception:
                    creation_date_str = str(creation_date)
            else:
                creation_date_str = 'N/A'

            if expiration_date:
                try:
                    expiration_date_str = datetime.fromtimestamp(expiration_date).strftime('%Y-%m-%d')
                except Exception:
                    expiration_date_str = str(expiration_date)
            else:
                expiration_date_str = 'N/A'

            # Determine if newly registered
            is_newly_registered = False
            if creation_date_obj:
                six_months_ago = datetime.now() - timedelta(days=180)
                is_newly_registered = creation_date_obj > six_months_ago

            # Create domain category
            domain_category = []
            if is_newly_registered:
                domain_category.append("Newly Registered Domain")
            if is_malicious:
                domain_category.append("Malicious")
            if malicious_count > 0:
                domain_category.append(f"Flagged by {malicious_count} vendors")
            domain_category_str = ', '.join(domain_category) if domain_category else 'Normal'

            # Calculate risk score
            risk_score = self.calculate_risk_score(
                reputation, malicious_count, suspicious_count, 
                categories_str, is_newly_registered
            )

            return {
                'Domain': domain,
                'Risk_Score': risk_score,
                'Reputation_Score': reputation,
                'Is_Malicious': is_malicious,
                'Malicious_Vendors': malicious_count,
                'Suspicious_Vendors': suspicious_count,
                'Harmless_Vendors': harmless_count,
                'Undetected_Vendors': undetected_count,
                'Total_Vendors': total_vendors,
                'Last_Scanned': last_analysis_date,
                'Tags': tags_str,
                'Categories': categories_str,
                'Domain_Category': domain_category_str,
                'Creation_Date': creation_date_str,
                'Expiration_Date': expiration_date_str,
                'Registrar': registrar,
                'Is_Newly_Registered': is_newly_registered,
                'WHOIS_Status': whois_status,
                'Query_Status': 'Success'
            }
        except Exception as e:
            return {
                'Domain': domain,
                'Risk_Score': 0,
                'Reputation_Score': 'N/A',
                'Is_Malicious': 'N/A',
                'Malicious_Vendors': 'N/A',
                'Suspicious_Vendors': 'N/A',
                'Harmless_Vendors': 'N/A',
                'Undetected_Vendors': 'N/A',
                'Total_Vendors': 'N/A',
                'Last_Scanned': 'N/A',
                'Tags': 'N/A',
                'Categories': 'N/A',
                'Domain_Category': 'N/A',
                'Creation_Date': 'N/A',
                'Expiration_Date': 'N/A',
                'Registrar': 'N/A',
                'Is_Newly_Registered': 'N/A',
                'WHOIS_Status': f'Error: {str(e)}',
                'Query_Status': f'Error: {str(e)}'
            }

    def review_domains(self, domains: List[str], progress_callback=None, status_callback=None) -> List[Dict]:
        results = []
        api_key_validated = False
        
        for i, domain in enumerate(domains):
            # Clean and defang domain
            original_domain = domain.strip()
            clean_domain = self.defang_domain(original_domain)
            
            if not clean_domain:
                continue
                
            if status_callback:
                if original_domain != clean_domain:
                    status_callback(f"Querying domain {i+1}/{len(domains)}: {original_domain} -> {clean_domain}")
                else:
                    status_callback(f"Querying domain {i+1}/{len(domains)}: {clean_domain}")
                    
            data = self.get_domain_info(clean_domain)
            
            # Check API key validity with first domain request
            if not api_key_validated and i == 0:
                if data is None:
                    # Check if it's an API key issue by making a test request
                    test_url = f"{self.base_url}/users/current"
                    test_response = requests.get(test_url, headers=self.headers)
                    if test_response.status_code == 401:
                        if status_callback:
                            status_callback("Error: Invalid API key detected")
                        raise ValueError("Invalid API key. Please check your VirusTotal API key.")
                api_key_validated = True
            
            if data:
                parsed_data = self.parse_domain_data(clean_domain, data)
                results.append(parsed_data)
            else:
                results.append({
                    'Domain': clean_domain,
                    'Risk_Score': 0,
                    'Reputation_Score': 'N/A',
                    'Is_Malicious': 'N/A',
                    'Malicious_Vendors': 'N/A',
                    'Suspicious_Vendors': 'N/A',
                    'Harmless_Vendors': 'N/A',
                    'Undetected_Vendors': 'N/A',
                    'Total_Vendors': 'N/A',
                    'Last_Scanned': 'N/A',
                    'Tags': 'N/A',
                    'Categories': 'N/A',
                    'Domain_Category': 'N/A',
                    'Creation_Date': 'N/A',
                    'Expiration_Date': 'N/A',
                    'Registrar': 'N/A',
                    'Is_Newly_Registered': 'N/A',
                    'WHOIS_Status': 'No data',
                    'Query_Status': 'Failed to query VirusTotal'
                })
            if progress_callback:
                progress_callback((i + 1) / len(domains) * 100)
            if i < len(domains) - 1:
                if status_callback:
                    status_callback(f"Waiting {self.rate_limit_delay} seconds before next query...")
                time.sleep(self.rate_limit_delay)
        return results

    def save_to_excel(self, results: List[Dict], output_file: str = "domain_review_results.xlsx"):
        if not results:
            return False
        try:
            df = pd.DataFrame(results)
            
            # Sort by risk score (highest risk first)
            df = df.sort_values('Risk_Score', ascending=False)
            
            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Domain_Review_Results', index=False)
                
                workbook = writer.book
                worksheet = writer.sheets['Domain_Review_Results']
                
                # Auto-adjust column widths
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
                
                # Apply heatmap formatting based on risk score
                from openpyxl.styles import PatternFill
                from openpyxl.formatting.rule import ColorScaleRule
                
                # Find the Risk_Score column
                risk_score_col = None
                for col_num, col in enumerate(df.columns, 1):
                    if col == 'Risk_Score':
                        risk_score_col = col_num
                        break
                
                if risk_score_col:
                    # Apply color scale to Risk_Score column
                    color_scale = ColorScaleRule(
                        start_type='min', start_color='00FF00',  # Green for low risk
                        mid_type='percentile', mid_value=50, mid_color='FFFF00',  # Yellow for medium risk
                        end_type='max', end_color='FF0000'  # Red for high risk
                    )
                    worksheet.conditional_formatting.add(
                        f'{chr(64 + risk_score_col)}2:{chr(64 + risk_score_col)}{len(df) + 1}',
                        color_scale
                    )
                    
                    # Apply row-based heatmap coloring
                    for row_idx, risk_score in enumerate(df['Risk_Score'], 2):
                        if risk_score > 0:
                            # Calculate color intensity based on updated risk score ranges
                            if risk_score > 70:
                                # High risk (>70) - Bright Red
                                fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
                            elif risk_score >= 50:
                                # Medium-high risk (50-70) - Orange
                                fill = PatternFill(start_color='FF6600', end_color='FF6600', fill_type='solid')
                            elif risk_score >= 30:
                                # Medium risk (30-50) - Yellow
                                fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                            elif risk_score >= 10:
                                # Low-medium risk (10-30) - Light yellow
                                fill = PatternFill(start_color='FFFACD', end_color='FFFACD', fill_type='solid')
                            else:
                                # Low risk (0-10) - Light green
                                fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
                            
                            # Apply fill to the entire row
                            for col in range(1, len(df.columns) + 1):
                                cell = worksheet.cell(row=row_idx, column=col)
                                cell.fill = fill
                
                # Add a summary sheet
                summary_data = {
                    'Metric': [
                        'Total Domains',
                        'High Risk (>70)',
                        'Medium-High Risk (50-70)',
                        'Medium Risk (30-50)',
                        'Low-Medium Risk (10-30)',
                        'Low Risk (0-10)',
                        'Malicious Domains',
                        'Newly Registered Domains',
                        'Average Risk Score'
                    ],
                    'Count': [
                        len(df),
                        len(df[df['Risk_Score'] > 70]),
                        len(df[(df['Risk_Score'] >= 50) & (df['Risk_Score'] <= 70)]),
                        len(df[(df['Risk_Score'] >= 30) & (df['Risk_Score'] < 50)]),
                        len(df[(df['Risk_Score'] >= 10) & (df['Risk_Score'] < 30)]),
                        len(df[df['Risk_Score'] < 10]),
                        len(df[df['Is_Malicious'] == True]),
                        len(df[df['Is_Newly_Registered'] == True]),
                        round(df['Risk_Score'].mean(), 2) if len(df) > 0 else 0
                    ]
                }
                
                summary_df = pd.DataFrame(summary_data)
                summary_df.to_excel(writer, sheet_name='Risk_Summary', index=False)
                
                # Format summary sheet
                summary_worksheet = writer.sheets['Risk_Summary']
                for column in summary_worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 30)
                    summary_worksheet.column_dimensions[column_letter].width = adjusted_width
            
            return True
        except Exception as e:
            print(f"Error saving to Excel: {e}")
            return False


def load_domains_from_file(filename: str) -> List[str]:
    try:
        with open(filename, 'r', encoding='utf-8') as file:
            domains = [line.strip() for line in file if line.strip()]
        return domains
    except Exception as e:
        return []

class DomainReviewerGUI:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("Domain Reviewer Tool")
        self.root.geometry("600x550")  # Increased height for new field
        self.root.resizable(False, False)
        
        # Set custom icon if available
        try:
            if os.path.exists("domain_reviewer_icon.ico"):
                self.root.iconbitmap("domain_reviewer_icon.ico")
        except Exception:
            pass  # Use default icon if custom icon fails to load
        
        self.center_window()
        self.api_key_var = tk.StringVar()
        self.account_type_var = tk.StringVar(value="personal")
        self.file_path_var = tk.StringVar()
        self.xsoar_case_var = tk.StringVar()
        self.domains = []
        self.setup_ui()
        
    def center_window(self):
        self.root.update_idletasks()
        width = self.root.winfo_width()
        height = self.root.winfo_height()
        x = (self.root.winfo_screenwidth() // 2) - (width // 2)
        y = (self.root.winfo_screenheight() // 2) - (height // 2)
        self.root.geometry(f'{width}x{height}+{x}+{y}')
        
    def validate_number(self, P):
        """Validate that input contains only numbers"""
        if P == "":  # Allow empty string
            return True
        return P.isdigit()
        
    def setup_ui(self):
        main_frame = ttk.Frame(self.root, padding="20")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        title_label = ttk.Label(main_frame, text="Domain Reviewer Tool", font=("Arial", 16, "bold"))
        title_label.grid(row=0, column=0, columnspan=2, pady=(0, 20))
        
        # XSOAR Case Number
        ttk.Label(main_frame, text="XSOAR Case Number:").grid(row=1, column=0, sticky=tk.W, pady=5)
        vcmd = (self.root.register(self.validate_number), '%P')
        xsoar_entry = ttk.Entry(main_frame, textvariable=self.xsoar_case_var, width=20, validate='key', validatecommand=vcmd)
        xsoar_entry.grid(row=1, column=1, sticky=(tk.W, tk.E), pady=5, padx=(10, 0))
        
        # API Key
        ttk.Label(main_frame, text="VirusTotal API Key:").grid(row=2, column=0, sticky=tk.W, pady=5)
        api_key_frame = ttk.Frame(main_frame)
        api_key_frame.grid(row=2, column=1, sticky=(tk.W, tk.E), pady=5, padx=(10, 0))
        
        api_key_entry = ttk.Entry(api_key_frame, textvariable=self.api_key_var, width=40, show="*")
        api_key_entry.grid(row=0, column=0, sticky=(tk.W, tk.E))
        
        test_api_button = ttk.Button(api_key_frame, text="Test Key", command=self.test_api_key)
        test_api_button.grid(row=0, column=1, padx=(5, 0))
        
        # Account Type
        ttk.Label(main_frame, text="Account Type:").grid(row=3, column=0, sticky=tk.W, pady=5)
        account_frame = ttk.Frame(main_frame)
        account_frame.grid(row=3, column=1, sticky=(tk.W, tk.E), pady=5, padx=(10, 0))
        
        personal_radio = ttk.Radiobutton(account_frame, text="Personal (15s delay, max 500 domains)", variable=self.account_type_var, value="personal")
        personal_radio.grid(row=0, column=0, sticky=tk.W)
        
        enterprise_radio = ttk.Radiobutton(account_frame, text="Enterprise (1s delay, unlimited)", variable=self.account_type_var, value="enterprise")
        enterprise_radio.grid(row=1, column=0, sticky=tk.W)
        
        # File Selection
        ttk.Label(main_frame, text="Domain List File:").grid(row=4, column=0, sticky=tk.W, pady=5)
        file_frame = ttk.Frame(main_frame)
        file_frame.grid(row=4, column=1, sticky=(tk.W, tk.E), pady=5, padx=(10, 0))
        
        file_entry = ttk.Entry(file_frame, textvariable=self.file_path_var, width=40)
        file_entry.grid(row=0, column=0, sticky=(tk.W, tk.E))
        
        browse_button = ttk.Button(file_frame, text="Browse", command=self.browse_file)
        browse_button.grid(row=0, column=1, padx=(5, 0))
        
        # Progress section
        progress_frame = ttk.LabelFrame(main_frame, text="Progress", padding="10")
        progress_frame.grid(row=5, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=20)
        
        self.progress_var = tk.StringVar(value="Ready to start")
        self.progress_label = ttk.Label(progress_frame, textvariable=self.progress_var)
        self.progress_label.grid(row=0, column=0, sticky=tk.W)
        
        self.progress_bar = ttk.Progressbar(progress_frame, mode='determinate')
        self.progress_bar.grid(row=1, column=0, sticky=(tk.W, tk.E), pady=(5, 0))
        
        # Buttons
        button_frame = ttk.Frame(main_frame)
        button_frame.grid(row=6, column=0, columnspan=2, pady=20)
        
        self.start_button = ttk.Button(button_frame, text="Start Review", command=self.start_review)
        self.start_button.grid(row=0, column=0, padx=(0, 10))
        
        exit_button = ttk.Button(button_frame, text="Exit", command=self.root.quit)
        exit_button.grid(row=0, column=1)
        
        # Configure grid weights
        main_frame.columnconfigure(1, weight=1)
        api_key_frame.columnconfigure(0, weight=1)
        file_frame.columnconfigure(0, weight=1)
        progress_frame.columnconfigure(0, weight=1)
        
    def browse_file(self):
        filename = filedialog.askopenfilename(
            title="Select Domain List File",
            filetypes=[("Text files", "*.txt"), ("All files", "*.*")]
        )
        if filename:
            self.file_path_var.set(filename)
            self.validate_file()
            
    def validate_file(self):
        file_path = self.file_path_var.get()
        if not file_path:
            return
        try:
            domains = load_domains_from_file(file_path)
            if not domains:
                messagebox.showerror("Error", "No domains found in the selected file.")
                return
            if self.account_type_var.get() == "personal" and len(domains) > 500:
                messagebox.showerror("Error", f"Personal accounts are limited to 500 domains. Found {len(domains)} domains.")
                return
            self.domains = domains
            messagebox.showinfo("Success", f"Loaded {len(domains)} domains from file.")
        except Exception as e:
            messagebox.showerror("Error", f"Error reading file: {str(e)}")
            
    def validate_inputs(self):
        # Check XSOAR Case Number
        if not self.xsoar_case_var.get().strip():
            messagebox.showerror("Error", "Please enter the XSOAR Case Number.")
            return False
            
        # Check API key
        if not self.api_key_var.get().strip():
            messagebox.showerror("Error", "Please enter your VirusTotal API key.")
            return False
            
        # Check file
        if not self.file_path_var.get().strip():
            messagebox.showerror("Error", "Please select a domain list file.")
            return False
            
        try:
            domains = load_domains_from_file(self.file_path_var.get())
            if not domains:
                messagebox.showerror("Error", "No domains found in the selected file.")
                return False
            if self.account_type_var.get() == "personal" and len(domains) > 500:
                messagebox.showerror("Error", f"Personal accounts are limited to 500 domains. Found {len(domains)} domains.")
                return False
            self.domains = domains
            return True
        except Exception as e:
            messagebox.showerror("Error", f"Error reading file: {str(e)}")
            return False

    def test_api_key(self):
        """
        Test the API key independently
        """
        if not self.api_key_var.get().strip():
            messagebox.showerror("Error", "Please enter your VirusTotal API key first.")
            return
            
        # Disable the test button temporarily
        for widget in self.root.winfo_children():
            if isinstance(widget, ttk.Frame):
                for child in widget.winfo_children():
                    if isinstance(child, ttk.Frame):
                        for grandchild in child.winfo_children():
                            if isinstance(grandchild, ttk.Button) and grandchild.cget('text') == "Test Key":
                                grandchild['state'] = 'disabled'
                                break
        
        def test_key():
            try:
                reviewer = VirusTotalDomainReviewer(
                    self.api_key_var.get().strip(),
                    self.account_type_var.get()
                )
                
                self.update_status("Testing API key...")
                is_valid = reviewer.verify_api_key()
                
                if is_valid:
                    self.update_status("API key is valid!")
                    messagebox.showinfo("Success", "API key is valid and working correctly!")
                else:
                    self.update_status("API key verification failed.")
                    messagebox.showerror("Error", "Invalid API key. Please check your VirusTotal API key.")
            except Exception as e:
                self.update_status("API key test failed.")
                messagebox.showerror("Error", f"Error testing API key: {str(e)}")
            finally:
                # Re-enable the test button
                for widget in self.root.winfo_children():
                    if isinstance(widget, ttk.Frame):
                        for child in widget.winfo_children():
                            if isinstance(child, ttk.Frame):
                                for grandchild in child.winfo_children():
                                    if isinstance(grandchild, ttk.Button) and grandchild.cget('text') == "Test Key":
                                        grandchild['state'] = 'normal'
                                        break
        
        thread = threading.Thread(target=test_key)
        thread.daemon = True
        thread.start()

    def verify_api_key_async(self):
        """
        Verify API key in a separate thread to avoid blocking the GUI
        """
        try:
            reviewer = VirusTotalDomainReviewer(
                self.api_key_var.get().strip(),
                self.account_type_var.get()
            )
            
            self.update_status("Verifying API key...")
            is_valid = reviewer.verify_api_key()
            
            if is_valid:
                self.update_status("API key verified successfully!")
                return True
            else:
                messagebox.showerror("Error", "Invalid API key. Please check your VirusTotal API key and try again.")
                self.update_status("API key verification failed.")
                return False
        except Exception as e:
            messagebox.showerror("Error", f"Error verifying API key: {str(e)}")
            self.update_status("API key verification failed.")
            return False
            
    def update_progress(self, value):
        self.progress_bar['value'] = value
        self.root.update_idletasks()
        
    def update_status(self, message):
        self.progress_var.set(message)
        self.root.update_idletasks()
        
    def start_review(self):
        if not self.validate_inputs():
            return
            
        self.start_button['state'] = 'disabled'
        
        # Start review directly - API key will be validated with first domain request
        thread = threading.Thread(target=self.run_review)
        thread.daemon = True
        thread.start()
        
    def run_review(self):
        try:
            reviewer = VirusTotalDomainReviewer(
                self.api_key_var.get().strip(),
                self.account_type_var.get()
            )
            self.update_status(f"Starting review of {len(self.domains)} domains...")
            self.update_progress(0)
            results = reviewer.review_domains(
                self.domains,
                progress_callback=self.update_progress,
                status_callback=self.update_status
            )
            if results:
                # Create filename with XSOAR case number
                case_number = self.xsoar_case_var.get().strip()
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                output_file = f"Case_{case_number}_domain_review_results_{timestamp}.xlsx"
                
                self.update_status("Saving results to Excel...")
                success = reviewer.save_to_excel(results, output_file)
                if success:
                    successful_queries = sum(1 for r in results if r['Query_Status'] == 'Success')
                    malicious_domains = sum(1 for r in results if r.get('Is_Malicious') == True)
                    newly_registered = sum(1 for r in results if r.get('Is_Newly_Registered') == True)
                    
                    # Calculate risk statistics
                    risk_scores = [r.get('Risk_Score', 0) for r in results if r.get('Risk_Score') is not None]
                    avg_risk = sum(risk_scores) / len(risk_scores) if risk_scores else 0
                    high_risk = sum(1 for score in risk_scores if score > 70)
                    
                    summary = f"Review completed!\n\n"
                    summary += f"XSOAR Case Number: {case_number}\n"
                    summary += f"Total domains: {len(self.domains)}\n"
                    summary += f"Successfully queried: {successful_queries}\n"
                    summary += f"Malicious domains found: {malicious_domains}\n"
                    summary += f"Newly registered domains: {newly_registered}\n"
                    summary += f"High risk domains (>70): {high_risk}\n"
                    summary += f"Average risk score: {avg_risk:.1f}\n"
                    summary += f"Results saved to: {output_file}"
                    
                    self.update_status("Review completed successfully!")
                    self.update_progress(100)
                    messagebox.showinfo("Review Complete", summary)
                else:
                    messagebox.showerror("Error", "Failed to save results to Excel file.")
            else:
                messagebox.showerror("Error", "No results to save.")
        except ValueError as e:
            # Handle API key validation error
            if "Invalid API key" in str(e):
                messagebox.showerror("Error", str(e))
                self.update_status("API key validation failed.")
            else:
                messagebox.showerror("Error", f"Validation error: {str(e)}")
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred during review: {str(e)}")
        finally:
            self.start_button['state'] = 'normal'
            
    def run(self):
        self.root.mainloop()

def main():
    app = DomainReviewerGUI()
    app.run()

if __name__ == "__main__":
    main() 